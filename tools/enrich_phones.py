#!/usr/bin/env python3
"""
Mac AI — Phone Number Enrichment Script
========================================
Reads the lead list Excel file, attempts to find phone numbers
by scraping each business's website, then exports an enriched CSV.

Usage:
  pip install openpyxl requests beautifulsoup4
  python3 enrich_phones.py

Output:
  enriched_leads.csv  — same data + Phone column filled where found
  enrichment_log.txt  — per-URL result log
"""

import openpyxl
import requests
from bs4 import BeautifulSoup
import re
import csv
import time
import os
import sys

# ── Config ────────────────────────────────────────────────────────────────────
INPUT_FILE  = os.path.expanduser(
    '~/Library/Application Support/Claude/local-agent-mode-sessions/'
    'd3a234f2-9241-4942-9e28-d6aa00579122/35e1ed45-3cfc-4860-a0a2-358126ce639f/'
    'agent/local_ditto_35e1ed45-3cfc-4860-a0a2-358126ce639f/outputs/'
    'Get_Mac_AI_Scotland_Leads_200.xlsx')
OUTPUT_CSV  = os.path.join(os.path.dirname(__file__), 'enriched_leads.csv')
LOG_FILE    = os.path.join(os.path.dirname(__file__), 'enrichment_log.txt')
DELAY_SEC   = 1.2   # polite delay between requests
TIMEOUT     = 8     # seconds per request

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/124.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'en-GB,en;q=0.9',
}

# UK phone regex — matches 01xxx, 02xxx, 03xxx, 07xxx, +44 variants
UK_PHONE_RE = re.compile(
    r'(?:(?:\+44|0044)[\s\-\.]?(?:\(0\))?[\s\-\.]?|0)'  # prefix
    r'(?:(?:1[0-9]{3}|2[0-9]|3[0-9]{2}|7[0-9]{3})'      # area/mobile
    r'[\s\-\.]?\d{3}[\s\-\.]?\d{3,4}'                     # subscriber
    r'|(?:800|808)[\s\-\.]?\d{3}[\s\-\.]?\d{4})'          # freephone
)

def clean_phone(raw: str) -> str:
    """Normalise a matched phone string."""
    num = re.sub(r'[\s\-\.\(\)]', '', raw).strip()
    if num.startswith('+44'):
        num = '0' + num[3:]
    return num

def find_phone_on_page(url: str):
    """Try to find a UK phone number on a given URL. Returns first found or None."""
    try:
        resp = requests.get(url, timeout=TIMEOUT, headers=HEADERS, allow_redirects=True)
        if resp.status_code >= 400:
            return None
        soup = BeautifulSoup(resp.text, 'html.parser')

        # 1. Look for tel: href links first (most reliable)
        for link in soup.find_all('a', href=re.compile(r'^tel:')):
            raw = link['href'].replace('tel:', '').strip()
            if raw and re.search(r'\d{7,}', raw):
                return clean_phone(raw)

        # 2. Look for schema.org telephone markup
        for tag in soup.find_all(itemprop='telephone'):
            txt = tag.get_text(strip=True)
            if txt and re.search(r'\d{7,}', txt):
                return clean_phone(txt)

        # 3. Scan visible text
        text = soup.get_text(separator=' ')
        matches = UK_PHONE_RE.findall(text)
        if matches:
            return clean_phone(matches[0])

    except requests.exceptions.Timeout:
        pass
    except requests.exceptions.ConnectionError:
        pass
    except Exception as e:
        pass
    return None


def find_phone_for_lead(website_url: str, business_name: str):
    """
    Try the main URL, then /contact and /contact-us pages.
    Returns (phone_or_None, status_string)
    """
    if not website_url or website_url.lower() in ['n/a', 'treatwell only', 'no', '']:
        return None, 'SKIP — no website'

    # Treatwell listing — can try scraping
    if 'treatwell' in website_url.lower():
        phone = find_phone_on_page(website_url if website_url.startswith('http')
                                    else 'https://' + website_url)
        return phone, ('FOUND via Treatwell' if phone else 'NOT FOUND — Treatwell listing')

    base = website_url if website_url.startswith('http') else 'https://' + website_url
    base = base.rstrip('/')

    pages_to_try = [base, base + '/contact', base + '/contact-us', base + '/about']

    for page in pages_to_try:
        phone = find_phone_on_page(page)
        if phone:
            return phone, f'FOUND on {page}'
        time.sleep(0.3)

    return None, f'NOT FOUND — tried {len(pages_to_try)} pages'


def main():
    print(f"\n{'='*60}")
    print('Mac AI — Phone Enrichment Script')
    print(f"{'='*60}")

    if not os.path.exists(INPUT_FILE):
        print(f'\n❌  Input file not found:\n    {INPUT_FILE}')
        print('\nUpdate INPUT_FILE path at the top of this script and retry.')
        sys.exit(1)

    wb  = openpyxl.load_workbook(INPUT_FILE)
    ws  = wb.active
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers  = [cell.value for cell in ws[1]]

    print(f'\n📂  Loaded: {os.path.basename(INPUT_FILE)}')
    print(f'    {len(all_rows)} leads, {len(headers)} columns')

    # Column indices from the Excel file
    COL = {h: i for i, h in enumerate(headers)}
    idx_name    = COL.get('Business Name', 0)
    idx_phone   = COL.get('Phone', 3)
    idx_website = COL.get('Website URL', 7)

    already_have = sum(1 for r in all_rows if r[idx_phone] and r[idx_phone] != 'N/A')
    to_enrich    = [r for r in all_rows if not r[idx_phone] or r[idx_phone] == 'N/A']
    print(f'\n    Already have phone : {already_have}')
    print(f'    Needs enrichment   : {len(to_enrich)}')

    print('\n🔍  Starting enrichment…')
    print('    (This will take a few minutes — being polite to servers)\n')

    results = []
    found_count = 0
    log_lines = []

    for i, row in enumerate(all_rows):
        name    = row[idx_name] or 'Unknown'
        phone   = row[idx_phone]
        website = row[idx_website] or ''
        row_mutable = list(row)

        if phone and phone != 'N/A':
            # Already has a number — keep it
            results.append(row_mutable)
            log_lines.append(f'[{i+1:03d}] {name:<40} KEPT existing: {phone}')
            continue

        enriched_phone, status = find_phone_for_lead(website, name)

        if enriched_phone:
            row_mutable[idx_phone] = enriched_phone
            found_count += 1
            marker = '✅'
        else:
            marker = '–'

        print(f'  {marker} [{i+1:03d}/{len(all_rows)}] {name[:40]:<40} {status}')
        log_lines.append(f'[{i+1:03d}] {name:<40} {status}')
        results.append(row_mutable)

        if website and website.lower() not in ['n/a', 'treatwell only', 'no', '']:
            time.sleep(DELAY_SEC)

    # Write CSV
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(results)

    # Write log
    with open(LOG_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(log_lines))

    print(f'\n{"="*60}')
    print(f'✅  Done!')
    print(f'    Found {found_count} new phone numbers')
    print(f'    Total with phone: {already_have + found_count} / {len(all_rows)}')
    print(f'\n    Output : {OUTPUT_CSV}')
    print(f'    Log    : {LOG_FILE}')
    print(f'{"="*60}\n')

    print('NEXT STEPS:')
    print('  1. Open enriched_leads.csv in Excel or Numbers')
    print('  2. For any still missing phones:')
    print('     → Check their Instagram bio (most aesthetics clinics list it there)')
    print('     → Search "[Business Name] Glasgow phone" in Google')
    print('     → Check their Treatwell page directly')
    print('  3. Import enriched_leads.csv into the CRM via the Import tab\n')


if __name__ == '__main__':
    main()
