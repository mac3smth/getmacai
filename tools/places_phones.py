#!/usr/bin/env python3
"""
Mac AI — Google Places Phone Enrichment
=========================================
Uses the Google Places API to find phone numbers for each clinic
directly from Google Business Profile (GMB) listings.

This will get ~85-90% coverage vs the ~19% from website scraping.

SETUP (one time — 5 minutes):
  1. Go to console.cloud.google.com
  2. Create a new project (or use existing)
  3. Search "Places API" → Enable it
  4. Go to Credentials → Create Credentials → API Key
  5. Paste your key into API_KEY below (or set env var GOOGLE_PLACES_KEY)
  6. pip install requests openpyxl
  7. python3 places_phones.py

Free quota: $200/month credit = ~14,000 Place Details calls
Cost for 200 clinics: approx $1.40 (well within free tier)

Output:
  places_enriched.csv   — lead list with phone numbers filled in
  places_log.txt        — per-business result log
"""

import os
import csv
import time
import requests
import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────
API_KEY    = os.environ.get('GOOGLE_PLACES_KEY', '')   # ← paste your key here or set env var
DELAY_SEC  = 0.5   # polite delay (Places API rate limit: 10 req/s, we use 2/s)
REGION     = 'GB'  # bias results to UK

INPUT_FILE = os.path.expanduser(
    '~/Library/Application Support/Claude/local-agent-mode-sessions/'
    'd3a234f2-9241-4942-9e28-d6aa00579122/35e1ed45-3cfc-4860-a0a2-358126ce639f/'
    'agent/local_ditto_35e1ed45-3cfc-4860-a0a2-358126ce639f/outputs/'
    'Get_Mac_AI_Scotland_Leads_200.xlsx'
)
OUTPUT_CSV = os.path.join(os.path.dirname(__file__), 'places_enriched.csv')
LOG_FILE   = os.path.join(os.path.dirname(__file__), 'places_log.txt')

PLACES_SEARCH_URL  = 'https://maps.googleapis.com/maps/api/place/textsearch/json'
PLACES_DETAILS_URL = 'https://maps.googleapis.com/maps/api/place/details/json'

# ── Google Places API calls ───────────────────────────────────────────────────

def find_place_id(business_name: str, area: str) -> str | None:
    """Text search to find the most likely Place ID for a business."""
    query = f'{business_name} aesthetics {area} Scotland'
    try:
        r = requests.get(PLACES_SEARCH_URL, params={
            'query':  query,
            'region': REGION,
            'key':    API_KEY,
        }, timeout=8)
        data = r.json()
        if data.get('status') == 'OK' and data.get('results'):
            return data['results'][0]['place_id']
    except Exception as e:
        print(f'    Search error: {e}')
    return None


def get_place_phone(place_id: str) -> str | None:
    """Fetch place details to get the formatted phone number."""
    try:
        r = requests.get(PLACES_DETAILS_URL, params={
            'place_id': place_id,
            'fields':   'name,formatted_phone_number,international_phone_number',
            'key':      API_KEY,
        }, timeout=8)
        data = r.json()
        if data.get('status') == 'OK':
            result = data.get('result', {})
            return result.get('formatted_phone_number') or result.get('international_phone_number')
    except Exception as e:
        print(f'    Details error: {e}')
    return None


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print(f'\n{"="*60}')
    print('Mac AI — Google Places Phone Enrichment')
    print(f'{"="*60}')

    if not API_KEY:
        print('\n❌  No API key set.')
        print('\n   Either:')
        print('   a) Paste your key into API_KEY at the top of this script')
        print('   b) Set the environment variable:')
        print('      export GOOGLE_PLACES_KEY="AIzaSy..."')
        print('      python3 places_phones.py')
        print('\n   To get a key: console.cloud.google.com → APIs → Places API → Credentials\n')
        return

    if not os.path.exists(INPUT_FILE):
        print(f'\n❌  Input file not found:\n    {INPUT_FILE}')
        return

    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    headers  = [cell.value for cell in ws[1]]
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))

    idx_name    = headers.index('Business Name')
    idx_area    = headers.index('Area')
    idx_phone   = headers.index('Phone')

    already_have = sum(1 for r in all_rows if r[idx_phone] and r[idx_phone] != 'N/A')
    to_enrich    = [r for r in all_rows if not r[idx_phone] or r[idx_phone] == 'N/A']

    print(f'\n📂  Loaded: {len(all_rows)} leads')
    print(f'    Already have phone : {already_have}')
    print(f'    Needs enrichment   : {len(to_enrich)}')
    print(f'\n🔍  Querying Google Places API...\n')

    results   = []
    found     = 0
    log_lines = []

    for i, row in enumerate(all_rows):
        name  = row[idx_name] or 'Unknown'
        area  = row[idx_area] or 'Scotland'
        phone = row[idx_phone]
        row_m = list(row)

        if phone and phone != 'N/A':
            results.append(row_m)
            log_lines.append(f'[{i+1:03d}] {name:<45} KEPT: {phone}')
            continue

        # Step 1: Find Place ID
        place_id = find_place_id(name, area)
        time.sleep(DELAY_SEC)

        if not place_id:
            results.append(row_m)
            log_lines.append(f'[{i+1:03d}] {name:<45} NOT FOUND in Places search')
            print(f'  – [{i+1:03d}/{len(all_rows)}] {name[:45]:<45} No Places result')
            continue

        # Step 2: Get phone from Place Details
        phone_num = get_place_phone(place_id)
        time.sleep(DELAY_SEC)

        if phone_num:
            row_m[idx_phone] = phone_num
            found += 1
            print(f'  ✅ [{i+1:03d}/{len(all_rows)}] {name[:45]:<45} {phone_num}')
            log_lines.append(f'[{i+1:03d}] {name:<45} FOUND: {phone_num}')
        else:
            print(f'  – [{i+1:03d}/{len(all_rows)}] {name[:45]:<45} On Maps but no phone listed')
            log_lines.append(f'[{i+1:03d}] {name:<45} On Maps, no phone')

        results.append(row_m)

    # Write CSV
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(results)

    with open(LOG_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(log_lines))

    total_with_phone = already_have + found
    print(f'\n{"="*60}')
    print(f'✅  Done!')
    print(f'    New phones found  : {found}')
    print(f'    Total with phone  : {total_with_phone} / {len(all_rows)} ({round(total_with_phone/len(all_rows)*100)}%)')
    print(f'\n    CSV → {OUTPUT_CSV}')
    print(f'    Log → {LOG_FILE}')
    print(f'{"="*60}')
    print('\nNEXT: Drag places_enriched.csv into the CRM Import tab.\n')


if __name__ == '__main__':
    main()
